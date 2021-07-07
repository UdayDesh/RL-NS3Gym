import random
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.chart import LineChart, Reference

class ExcelChannel:
    """
    Provides a set of channel-signal-fill pattern to be followed
    """
    def shiftAppend(self,channels):
        k=[1]
        for j in range(1,channels): k.append(0)
        p=[]
        for i in range(channels):
            a = i % len(k)
            p.append(k[-a:] + k[:-a])
        return p
    """
    Creates a set of Channel numbers for document purpose e.g. C0,C1....etc
    """
    def createSet(self,channels):
        k=['']
        for j in range(channels): k.append('C'+str(j))
        return k
    """
    Initial method for documentation
    """
    def __init__(self,channels=4):
        self.trials = []
        self.pattern = self.shiftAppend(channels)
        self.filename = self.getFileName()
        self.channels = channels
        self.chnlSet = self.createSet(channels)
        self.wb = Workbook()
        self.iteration = self.wb.create_sheet("ITERATIONS")
        self.charts = self.wb.create_sheet("CHARTS")
        self.cumulative = self.wb.create_sheet("CUMULATIVE")

    """
    File Name is auto-generated from the current date based on Julian calendar <DAY><DD><MMM><YYYY>
    e.g. 24JUN2021_12:34:56
    """
    def getFileName(self):
        import time
        return 'ChannelTrace_'+time.strftime("%d%b%Y_%H-%M-%S")+'.xlsx'

    """
    Checks for Iterations & updates the XLS with Iteration-wise report
    """
    def inputIterations(self,trials):
        iter_count=len(trials)
        for trial in range(iter_count):
            rew_count=len(trials[trial])
            trial_pattern={0:self.chnlSet}
            for i in range(0,rew_count):
                temp = []
                temp.append('T'+str(i))
                for j in self.pattern[i%self.channels]: 
                    if(j==1 and trials[trial][i] < 0):j=-1
                    temp.append(j)
                trial_pattern[i+1]=temp
            
            self.iteration.cell(row=1+8*trial, column=1).value ="Iteration "+str(trial+1)
            for col in range(len(trial_pattern)):
                row_start=2+8*trial
                curr_col = trial_pattern[col]
                for i in range(len(curr_col)):
                    if(curr_col[i]==0):
                        continue
                    elif(curr_col[i]==1):
                        self.iteration.cell(row=row_start+i, column=col+1).fill = PatternFill("solid", fgColor="990000")
                    elif(curr_col[i] == -1):
                        #INTERFERENCE WITH PREDICTED VALUE
                        self.iteration.cell(row=row_start+i, column=col+1).fill = PatternFill("solid", fgColor="DDDDDD")
                    else:
                        self.iteration.cell(row=row_start+i, column=col+1).value = trial_pattern[col][i]
            self.wb.save(filename=self.filename)

    """
    Creates line chart for given trials data based on success rate
    """
    def inputChart(self,trials):
        count=1
        data=[['Iteration','Rate']]
        for iter in trials:
            rewards = sum(iter)
            total = len(iter)
            if total > 0:
                data.append([count,rewards*100/total])
            else:
                data.append([count,0])
            count = count + 1
        #print(count)
        #Dump this analysis in Charts sheet
        for d in data: self.charts.append(d)
        data = Reference(self.charts, min_col=2,max_col=2, min_row=1, max_row=count)
        c1 = LineChart()
        c1.title = "Iteration Progress"
        c1.style = 13
        c1.y_axis.title = 'Success%'
        c1.x_axis.title = 'Iteration#'
        c1.add_data(data, titles_from_data=True)
        
        s1 = c1.series[0]
        s1.marker.symbol = "triangle"
        s1.marker.graphicalProperties.solidFill = "DDDDDD" # Marker filling
        s1.marker.graphicalProperties.line.solidFill = "FF0000" # Marker outline
        #s1.graphicalProperties.line.noFill = True
        iters = Reference(self.charts, min_col=1, min_row=2, max_row=count)
        c1.set_categories(iters)

        self.charts.add_chart(c1, "E3")
        self.wb.save(filename=self.filename)

    """
    Creates line chart for given trials data based on success rate
    """
    def inputCumulative(self,trials):
        data=[['Iteration','Rate']]
        for i in range(len(trials)): data.append([i+1,trials[i]])
        count = len(trials)
        #print(count)
        #Dump this analysis in Cumulative sheet
        for d in data: self.cumulative.append(d)
        data = Reference(self.cumulative, min_col=2,max_col=2, min_row=1, max_row=count)
        c1 = LineChart()
        c1.title = "Iteration Cumulative Rewards"
        c1.style = 13
        c1.y_axis.title = 'Cumulative Reward%'
        c1.x_axis.title = 'Iteration#'
        c1.add_data(data, titles_from_data=True)
        
        s1 = c1.series[0]
        s1.marker.symbol = "circle"
        s1.marker.graphicalProperties.solidFill = "DDDDDD" # Marker filling
        s1.marker.graphicalProperties.line.solidFill = "FF0000" # Marker outline
        #s1.graphicalProperties.line.noFill = True
        iters = Reference(self.cumulative, min_col=1, min_row=2, max_row=count)
        c1.set_categories(iters)

        self.cumulative.add_chart(c1, "E3")
        self.wb.save(filename=self.filename)

if __name__ == "__main__":
    ec = ExcelChannel()
    # Generate trials data
    iterations = random.randint(1, 200)# Pick a random number between 1 and 200.
    rewards=[]
    for i in range(iterations):
        t = random.randint(1, 50)
        t_success = random.randint(26,50)#Successful is more
        t_failure = 50 - t_success
        k = [1 for i in range(t_success)] + [-1 for j in range(t_failure)]
        random.shuffle(k)#Mix of 1 & -1s
        rewards.append(k)
    
    #rewards = [[1,1,1,1,-1,1,-1,-1,-1,1,1,1,1],[1,1,1,1,-1,1,-1,-1,-1,1,1],[1,1,1,1,-1,1,-1],[-1,-1,-1,-1,-1,1,-1,-1,-1,1,1,1,1]]
    ec.inputIterations(rewards)
    ec.inputChart(rewards)
    
